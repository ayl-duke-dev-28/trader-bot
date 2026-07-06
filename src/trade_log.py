"""Append-only Excel activity log for trades.

One row per broker-facing action (submitted buy/sell, stop-loss close,
skipped intent, dry-run intent, failure) so the user can review *why*
each trade was made after the fact.
"""
from __future__ import annotations

import logging
import threading
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.config import Config, ROOT

log = logging.getLogger(__name__)

SHEET_NAME = "trades"
COLUMNS: tuple[str, ...] = (
    "timestamp",
    "mode",
    "action",
    "symbol",
    "qty",
    "price",
    "target_dollars",
    "score",
    "reason",
    "order_id",
)


@dataclass(frozen=True)
class TradeLogEntry:
    action: str          # BUY | SELL | STOP | SKIP | DRY | FAIL
    symbol: str
    mode: str            # paper | live | dry
    qty: float = 0.0
    price: float = 0.0
    target_dollars: float = 0.0
    score: float | None = None
    reason: str = ""
    order_id: str = ""


class TradeLogger:
    """Thread-safe append-to-xlsx logger."""

    def __init__(self, path: Path):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()
        if not self.path.exists():
            self._create()

    def _create(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(list(COLUMNS))
        wb.save(self.path)

    def log(self, entry: TradeLogEntry) -> None:
        row = (
            datetime.now().isoformat(timespec="seconds"),
            entry.mode,
            entry.action,
            entry.symbol,
            round(float(entry.qty), 6),
            round(float(entry.price), 4),
            round(float(entry.target_dollars), 2),
            "" if entry.score is None else round(float(entry.score), 4),
            entry.reason,
            entry.order_id,
        )
        try:
            with self._lock:
                wb = load_workbook(self.path)
                ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
                ws.append(list(row))
                wb.save(self.path)
        except Exception as e:
            log.warning("trade log write failed for %s %s: %s", entry.action, entry.symbol, e)


def trade_logger_from_config(cfg: Config) -> TradeLogger:
    path = Path(cfg.get("logging", "trades_file", default="logs/trades.xlsx"))
    if not path.is_absolute():
        path = ROOT / path
    return TradeLogger(path)
